from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side


class UserService:

    def download_user_info(self, data):
        columns = {
            'ID': 1,
            'Email': 2,
            'Telegram': 3,
            'Instagram': 4,
            'Phone': 5,
        }

        wb = Workbook()
        ws = wb.active

        # Headers
        for key, value in columns.items():
            ws.cell(row=1, column=value).value = key

            # Styles
            ws.cell(row=1, column=value).font = Font(
                color='FFFFFF'
            )
            ws.cell(row=1, column=value).fill = PatternFill(
                start_color='6495ED',
                fill_type='solid'
            )
            ws.cell(row=1, column=value).border = Border(
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )

        # Data
        for ui in data:
            row_data = [
                ui.get('id'),
                ui.get('email'),
                ui.get('telegram'),
                ui.get('instagram'),
                ui.get('phone')
            ]
            ws.append(row_data)

        # Filter for columns
        ws.auto_filter.ref = ws.dimensions
        # Adjust column size
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length
        wb.close()

        return wb
